import os
import json
from openpyxl import load_workbook
import shutil

def export_dataset_and_labels_to_json(excel_file, dataset_folder, label_file):
    try:
        # Create dataset folder if it doesn't exist
        if not os.path.exists(dataset_folder):
            os.makedirs(dataset_folder)
    except PermissionError as e:
        print(f"PermissionError: Unable to create or access the folder '{dataset_folder}'. Please check your permissions.")
        return

    data = []  # List to store dataset and labels as JSON objects

    try:
        # Load Excel file
        wb = load_workbook(excel_file)
        ws = wb.active

        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            image_name, ocr_value = row[:2]  # Read "Image Name" and "Value" columns

            # Check if image_name or ocr_value is None
            if image_name is None or ocr_value is None:
                print(f"Skipping invalid row: {row}")
                continue

            # Convert to strings and validate if both are non-empty
            image_name = str(image_name).strip()
            ocr_value = str(ocr_value).strip()
            if not image_name or not ocr_value:
                print(f"Skipping invalid row: {row}")
                continue

            # Source image path in the dataset/images folder
            source_image_path = os.path.join(dataset_folder, "images", image_name)

            # Validate if the image file exists
            if not os.path.isfile(source_image_path):
                print(f"File not found: {source_image_path}. Skipping...")
                continue

            # Append the data as a dictionary to the list
            data.append({"image_name": image_name, "ocr_value": ocr_value})
            print(f'Exported: {image_name} -> Label: {ocr_value}')

        # Write the data to a JSON file
        with open(label_file, "w", encoding="utf-8") as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)
            print(f"Data successfully exported to {label_file}")
    except PermissionError as e:
        print(f"PermissionError: Unable to write to the file '{label_file}'. Please check your permissions.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    excel_file = r"./dataset_labels.xlsx"  # Input Excel file
    dataset_folder = r"./dataset"  # Dataset folder
    label_file = r"./dataset/labels.json"  # Output JSON file
    export_dataset_and_labels_to_json(excel_file, dataset_folder, label_file)
