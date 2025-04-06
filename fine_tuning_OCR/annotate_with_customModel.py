import os
import torch
from PIL import Image as PILImage
from torchvision import transforms
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Define the custom OCR model (replace with your actual model definition)
class CustomOCRModel(torch.nn.Module):
    def __init__(self, num_classes):
        super(CustomOCRModel, self).__init__()
        self.conv = torch.nn.Sequential(
            torch.nn.Conv2d(1, 64, kernel_size=3, stride=1, padding=1),
            torch.nn.ReLU(),
            torch.nn.MaxPool2d(kernel_size=2, stride=2)
        )
        self.rnn = torch.nn.LSTM(64 * 16, 128, bidirectional=True, batch_first=True)
        self.fc = torch.nn.Linear(128 * 2, num_classes)  # Bidirectional LSTM doubles the hidden size

    def forward(self, x):
        x = self.conv(x)  # Shape: (N, C, H, W)
        x = x.permute(0, 3, 1, 2)  # Shape: (N, W, C, H)
        x = x.flatten(2)  # Shape: (N, W, C*H)
        x, _ = self.rnn(x)  # Shape: (N, W, 2*hidden_size)
        x = self.fc(x)  # Shape: (N, W, num_classes)
        return x

# Load the custom model
def load_model(model_path, num_classes):
    model = CustomOCRModel(num_classes=num_classes)
    state_dict = torch.load(model_path, map_location=torch.device('cpu'))

    # Check for size mismatches and adjust if necessary
    model_state_dict = model.state_dict()
    for key in state_dict.keys():
        if key in model_state_dict and state_dict[key].size() != model_state_dict[key].size():
            print(f"Skipping loading parameter: {key} due to size mismatch.")
            state_dict[key] = model_state_dict[key]  # Use the default initialized weights for mismatched layers

    model.load_state_dict(state_dict, strict=False)
    model.eval()
    return model

# Perform OCR using the custom model
def extract_text_with_custom_model(image_path, model, transform, idx_to_char):
    try:
        image = PILImage.open(image_path).convert("L")  # Convert to grayscale
        image = transform(image).unsqueeze(0)  # Add batch dimension
        with torch.no_grad():
            outputs = model(image)  # Shape: (T, N, C)
            outputs = outputs.permute(1, 0, 2)  # Shape: (N, T, C)
            predicted_indices = torch.argmax(outputs, dim=2).squeeze(0)  # Get the most likely indices
            predicted_text = "".join([idx_to_char[idx.item()] for idx in predicted_indices])
        return predicted_text.strip()
    except Exception as e:
        print(f"Error processing {image_path}: {e}")
        return "Error solving OCR"

def process_images_in_folder(folder_path, output_excel, model, transform, idx_to_char):
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"
    ws.append(["Image Path", "Detected Measurement Value", "Image Preview"])  # Updated header row

    # Set column width and row height for image cells
    ws.column_dimensions["C"].width = 25  # Adjust column width
    row_height = 75  # Adjust row height

    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
                file_path = os.path.join(root, file)
                print(f"Processing file: {file_path}")
                # Solve OCR using the custom model
                detected_text = extract_text_with_custom_model(file_path, model, transform, idx_to_char)
                print(f"Detected measurement value: {detected_text}")
                ws.append([file_path, detected_text])  # Add data to Excel

                # Insert image into Excel
                img = Image(file_path)
                ws.row_dimensions[ws.max_row].height = row_height  # Set row height
                img.width = 100  # Resize image width to fit cell
                img.height = 75  # Resize image height to fit cell
                ws.add_image(img, f"C{ws.max_row}")  # Insert image into column C of the current row

    wb.save(output_excel)
    print(f"Results saved to {output_excel}")

if __name__ == "__main__":
    folder_path = r"C:\Users\ndt31\Desktop\DFM-SMT1-CSA\PREDICT_OUTPUT\Cropped_Images"
    output_excel = "ocr_results.xlsx"  # Output Excel file name
    model_path = r"C:\Users\ndt31\Desktop\DFM-SMT1-CSA\project\fine_tuned_model.pth"  # Path to the fine-tuned model

    # Define character set and mapping
    charset = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    idx_to_char = {idx: char for idx, char in enumerate(charset)}

    # Load the model
    model = load_model(model_path, num_classes=len(charset))

    # Define transformations
    transform = transforms.Compose([
        transforms.Resize((32, 128)),  # Resize images to a fixed size
        transforms.ToTensor(),
        transforms.Normalize((0.5,), (0.5,))  # Normalize to [-1, 1]
    ])

    if os.path.exists(folder_path):
        process_images_in_folder(folder_path, output_excel, model, transform, idx_to_char)
    else:
        print("The specified folder does not exist.")
