import os
import shutil
import easyocr  # Offline OCR library
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Initialize EasyOCR reader
reader = easyocr.Reader(['en'])  # Specify language(s) as needed

def create_dataset_and_excel(parent_folder, dataset_folder, output_excel):
    # Create dataset/images folder if it doesn't exist
    images_folder = os.path.join(dataset_folder, "images")
    os.makedirs(images_folder, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset"
    ws.append(["Image Name", "Value", "Image"])  # Updated header row

    ws.column_dimensions["C"].width = 25  # Adjust column width
    row_height = 75  # Adjust row height

    for root, _, files in os.walk(parent_folder):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
                abs_path = os.path.join(root, file)
                relative_path = os.path.relpath(root, parent_folder).replace(os.sep, "_")
                image_name = f"{relative_path}_{file}" if relative_path != "." else file
                dataset_image_path = os.path.join(images_folder, image_name)

                # Copy image to dataset/images folder
                shutil.copy2(abs_path, dataset_image_path)

                # Perform OCR using EasyOCR
                try:
                    result = reader.readtext(abs_path, detail=0)  # Extract text without details
                    ocr_value = " ".join(result).strip()  # Combine detected text
                except Exception as e:
                    print(f"Error processing {abs_path}: {e}")
                    ocr_value = "Error solving OCR"

                # Append data to Excel
                ws.append([image_name, ocr_value])  # Add image name and OCR value
                img = Image(dataset_image_path)
                ws.row_dimensions[ws.max_row].height = row_height  # Set row height
                img.width = 100  # Resize image width to fit cell
                img.height = 60  # Resize image height to fit cell
                ws.add_image(img, f"C{ws.max_row}")  # Insert image into column C of the current row

                print(f"Processed: {image_name} -> OCR: {ocr_value}")

    # Save Excel file
    wb.save(output_excel)
    print(f"Dataset and OCR results saved to {output_excel}")

if __name__ == "__main__":
    parent_folder = r"C:\Users\ndt31\Desktop\DFM-SMT1-CSA\PREDICT_OUTPUT\Cropped_Images"  # Parent folder containing images
    dataset_folder = r"C:\Users\ndt31\Desktop\DFM-SMT1-CSA\project\dataset"  # Dataset folder
    output_excel = r"C:\Users\ndt31\Desktop\DFM-SMT1-CSA\project\dataset_labels.xlsx"  # Output Excel file
    create_dataset_and_excel(parent_folder, dataset_folder, output_excel)
