import os
import cv2
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from PIL import Image, ImageEnhance, ExifTags, ImageTk  # Import ExifTags for handling image orientation
import tkinter as tk
from tkinter import messagebox, Toplevel, Label, Entry, Button, filedialog, Canvas  # Import Canvas for image display
import subprocess
from openpyxl.styles import Alignment  # Import Alignment for cell formatting
from tkinter.colorchooser import askcolor  # Import color chooser for selecting text color
import numpy as np  # Import numpy for array operations
import sys  # Import sys for restarting the script
import torch
from torchvision import transforms
import torch.nn as nn
from difflib import SequenceMatcher  # Import for similarity calculation

cropped_images_folder = "./Predict_Output/Cropped_Images"
output_excel_path = "./excel_output/results.xlsx"
margin_value = 0

custom_model_folder = None  # Remove custom model folder functionality

margin_top = 0
margin_bottom = 0
margin_left = 0
margin_right = 0

model_storage_directory = "C:/Users/ndt31/Desktop/python-3.13.2-embed-amd64/YOLO_OCR/OCR_custom_model"  # Directory for the fine-tuned model

wb = Workbook()
ws = wb.active
ws.title = "OCR Results"

headers = ["Path", "Score", "OCR Result", "Cropped Img"]
ws.append(headers)

debug_folder = "./debug_cropped_images"
os.makedirs(debug_folder, exist_ok=True)

selected_color = None  # Global variable to store the selected text color

color_range = {"hue": 10, "saturation": 50, "value": 50}  # Default range for HSV adjustments

tolerance_value = 50  # Default tolerance value for color filtering

apply_filter = True  # Global variable to track whether the filter is applied

model_path = r"./OCR_custom_model/model1.pth"  # Default model path

# Define the CRNN model
class CRNN(nn.Module):
    def __init__(self, num_classes):
        super(CRNN, self).__init__()
        self.conv = nn.Sequential(
            nn.Conv2d(1, 64, kernel_size=3, stride=1, padding=1),
            nn.ReLU(),
            nn.MaxPool2d(kernel_size=2, stride=2),
            nn.Dropout(0.3)
        )
        self.rnn = nn.LSTM(64 * 16, 128, bidirectional=True, batch_first=True)
        self.fc = nn.Linear(128 * 2, num_classes)
        self.output_length = 32

    def forward(self, x):
        x = self.conv(x)
        x = x.permute(0, 3, 1, 2)
        x = x.flatten(2)
        x, _ = self.rnn(x)
        x = self.fc(x)
        return x.permute(1, 0, 2)

# Function to load the model
def load_model(model_path, char_to_idx):
    num_classes = len(char_to_idx) + 1  # Add 1 for the blank token
    model = CRNN(num_classes=num_classes)
    model.load_state_dict(torch.load(model_path, map_location=torch.device("cpu")))
    model.eval()
    return model

# Function to perform OCR on a single image
def ocr_image(model, image_path, transform, idx_to_char, blank_idx):
    """
    Perform OCR on a single image and calculate confidence scores.
    """
    image = Image.open(image_path).convert("L")
    image = transform(image).unsqueeze(0)  # Add batch dimension
    with torch.no_grad():
        outputs = model(image)
        outputs = outputs.log_softmax(2)  # Shape: (T, N, C)
        predicted_indices = torch.argmax(outputs, dim=2).squeeze(1)  # Shape: (T,)
        confidences = torch.exp(outputs.max(dim=2).values).squeeze(1)  # Extract confidence scores

        cleaned_pred = []
        confidence_scores = []
        prev_idx = None
        for idx, confidence in zip(predicted_indices, confidences):
            idx = idx.item()
            if idx != prev_idx and idx != blank_idx:
                if idx_to_char[idx] != "<UNK>":
                    cleaned_pred.append(idx_to_char[idx])
                    confidence_scores.append(confidence.item())
            prev_idx = idx

        # Calculate average confidence score
        avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0
        return "".join(cleaned_pred), avg_confidence

# Initialize the CRNN model and transformations
charset = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789.,!?;:'\"()[]{}<>@#$%^&*-_+=/\\|~` "
unk_token = "<UNK>"
char_to_idx = {char: idx for idx, char in enumerate(charset)}
char_to_idx[unk_token] = len(char_to_idx)
blank_idx = len(char_to_idx)
idx_to_char = {idx: char for char, idx in char_to_idx.items()}

transform = transforms.Compose([
    transforms.Resize((32, 128)),
    transforms.ToTensor(),
    transforms.Normalize((0.5,), (0.5,))
])

if not os.path.exists(model_path):
    raise FileNotFoundError(f"Model file not found at {model_path}")
model = load_model(model_path, char_to_idx)

def crop_image(image_path, top=0, bottom=0, left=0, right=0):
    """
    Crop the image with specified margins for top, bottom, left, and right.
    """
    img = Image.open(image_path)
    width, height = img.size
    cropped_img = img.crop((left, top, width - right, height - bottom))
    return cropped_img

def correct_image_orientation(image):
    """
    Correct the orientation of an image using its EXIF metadata.
    """
    try:
        exif = image._getexif()
        if (exif is not None):
            for tag, value in ExifTags.TAGS.items():
                if (value == 'Orientation'):
                    orientation = exif.get(tag)
                    if (orientation == 3):  # Rotate 180 degrees
                        image = image.rotate(180, expand=True)
                    elif (orientation == 6):  # Rotate 270 degrees (clockwise)
                        image = image.rotate(270, expand=True)
                    elif (orientation == 8):  # Rotate 90 degrees (counterclockwise)
                        image = image.rotate(90, expand=True)
    except AttributeError:
        pass  # If no EXIF data, skip correction
    return image

def isolate_selected_color(image, selected_color, tolerance=None):
    """
    Retain only the selected color in the image and remove all other colors.
    Allow a specified RGB tolerance for color matching.
    """
    if not apply_filter:  # Skip filtering if apply_filter is False
        print("Filter is disabled. Returning the original image.")
        return image

    if selected_color is None:
        print("No color selected. Returning the original image.")
        return image  # If no color is selected, return the original image

    if tolerance is None:
        tolerance = tolerance_value  # Use the global tolerance value if not provided

    # Ensure the image is in RGB mode
    if image.mode != "RGB":
        print("Converting image to RGB mode.")
        image = image.convert("RGB")

    # Convert the image to OpenCV format (BGR)
    image_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)

    # Extract RGB values from the selected color
    r, g, b = selected_color

    # Define the lower and upper bounds for the selected color with tolerance
    lower_bound = np.array([max(0, b - tolerance), max(0, g - tolerance), max(0, r - tolerance)], dtype=np.uint8)
    upper_bound = np.array([min(255, b + tolerance), min(255, g + tolerance), min(255, r + tolerance)], dtype=np.uint8)

    # Debug: Print the bounds
    print(f"Selected Color: R={r}, G={g}, B={b}")
    print(f"Color Bounds: Lower={lower_bound}, Upper={upper_bound}")

    # Create a mask for the selected color
    mask = cv2.inRange(image_cv, lower_bound, upper_bound)

    # Debug: Save the mask for verification
    cv2.imwrite("debug_mask.png", mask)
    print("Saved debug mask as 'debug_mask.png'.")

    # Apply the mask to the original image
    result = cv2.bitwise_and(image_cv, image_cv, mask=mask)

    # Convert back to PIL Image
    result_image = Image.fromarray(cv2.cvtColor(result, cv2.COLOR_BGR2RGB))
    return result_image

def preprocess_image(image_path):
    """
    Preprocess the image by cropping and applying filters.
    """
    cropped_img = crop_image(image_path, margin_top, margin_bottom, margin_left, margin_right)
    cropped_img = correct_image_orientation(cropped_img)
    filtered_img = isolate_selected_color(cropped_img, selected_color)
    return filtered_img

def add_to_excel(relative_path, ocr_results, img_path, allowed_chars):
    cell_width = 150  # Increased cell width
    cell_height = 80  # Increased cell height

    if not ocr_results:
        row = [relative_path, "", "", ""]  # Add empty OCR results
        ws.append(row)

        cropped_img = Image.open(img_path)
        cropped_img.thumbnail((cell_width, cell_height))  # Resize image to fit in cell
        img_bytes = BytesIO()
        cropped_img.save(img_bytes, format="PNG")
        img_bytes.seek(0)
        excel_img = ExcelImage(img_bytes)
        img_col = chr(65 + len(row) - 1)  # Adjust for the last column (cropped_image)
        img_cell = f"{img_col}{ws.max_row}"
        ws.add_image(excel_img, img_cell)

        # Adjust row height and align content
        ws.row_dimensions[ws.max_row].height = cell_height * 0.75
        for col in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
    else:
        for bbox, text, score in ocr_results:
            filtered_text = ''.join(c for c in text if c in allowed_chars)
            if not filtered_text:
                continue

            row = [relative_path, score, filtered_text, ""]  # Adjusted order: path, score, OCR_result, cropped_image
            ws.append(row)

            cropped_img = Image.open(img_path)
            cropped_img.thumbnail((cell_width, cell_height))  # Resize image to fit in cell
            img_bytes = BytesIO()
            cropped_img.save(img_bytes, format="PNG")
            img_bytes.seek(0)
            excel_img = ExcelImage(img_bytes)
            img_col = chr(65 + len(row) - 1)  # Adjust for the last column (cropped_image)
            img_cell = f"{img_col}{ws.max_row}"
            ws.add_image(excel_img, img_cell)

            # Adjust row height and align content
            ws.row_dimensions[ws.max_row].height = cell_height * 0.75
            for col in range(1, len(row) + 1):
                ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

def restart_session():
    """
    Restart the current Python script.
    """
    print("Restarting the session...")
    os.execv(sys.executable, ['python'] + sys.argv)

def calculate_similarity(actual_text, predicted_text):
    """
    Calculate the similarity score between the actual text and the predicted text.
    """
    return SequenceMatcher(None, actual_text, predicted_text).ratio()

def process_images(flip_image, allowed_chars):
    # Ensure the output directory exists
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

    for root, _, files in os.walk(cropped_images_folder):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                img_path = os.path.join(root, file)
                relative_path = os.path.relpath(img_path, cropped_images_folder)

                print(f"Processing image: {relative_path}")

                try:
                    preprocessed_img = preprocess_image(img_path)
                    debug_image_path = os.path.join(debug_folder, os.path.basename(img_path))
                    preprocessed_img.save(debug_image_path)
                    print(f"Saved preprocessed debug image: {debug_image_path}")
                except Exception as e:
                    print(f"Error preprocessing {img_path}: {e}")
                    continue

                try:
                    ocr_text, score = ocr_image(model, debug_image_path, transform, idx_to_char, blank_idx)
                    print(f"OCR result for {relative_path}: Value='{ocr_text}', Score={score:.2f}")
                    ocr_results = [(None, ocr_text, score)]  # Mock bbox and use confidence score
                except Exception as e:
                    print(f"Error processing {debug_image_path}: {e}")
                    continue

                add_to_excel(relative_path, ocr_results, debug_image_path, allowed_chars)

    # Save the workbook with a robust method
    try:
        temp_output_path = output_excel_path + ".tmp"
        wb.save(temp_output_path)  # Save to a temporary file first
        os.replace(temp_output_path, output_excel_path)  # Atomically replace the old file
        print(f"Excel file saved to: {output_excel_path}")
        restart_session()  # Restart the session after saving
    except Exception as e:
        print(f"Error saving Excel file: {e}")

def open_sample_image():
    """
    Open a file dialog to select a sample image for color selection.
    Display the image in a new window with zooming, panning, and color picking capabilities.
    """
    file_path = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg")])
    if not file_path:
        return

    sample_image = Image.open(file_path)
    sample_image = sample_image.convert("RGB")  # Ensure the image is in RGB mode

    def on_click(event):
        """
        Handle mouse click on the image to pick a color.
        """
        global selected_color
        x, y = int(canvas.canvasx(event.x) / scale), int(canvas.canvasy(event.y) / scale)  # Adjust coordinates based on zoom scale
        selected_color = sample_image.getpixel((x, y))
        print(f"Selected color: {selected_color}")
        color_label.config(text=f"Selected Color: {selected_color}", bg=f"#{selected_color[0]:02x}{selected_color[1]:02x}{selected_color[2]:02x}")

    def update_tolerance(value):
        """
        Update the global tolerance value based on the slider.
        """
        global tolerance_value
        tolerance_value = int(value)
        print(f"Updated tolerance value to: {tolerance_value}")

    def apply_zoom():
        """
        Apply zoom based on the current scale.
        """
        nonlocal tk_image
        resized_image = sample_image.resize((int(original_width * scale), int(original_height * scale)), Image.Resampling.LANCZOS)
        tk_image = ImageTk.PhotoImage(resized_image)
        canvas.itemconfig(image_id, image=tk_image)
        canvas.configure(scrollregion=canvas.bbox("all"))

    def zoom(event):
        """
        Handle zooming in and out of the image using the mouse wheel.
        """
        nonlocal scale
        if event.delta > 0:  # Zoom in
            scale *= 1.1
        elif event.delta < 0:  # Zoom out
            scale /= 1.1
        apply_zoom()

    def start_pan(event):
        """
        Start panning the image.
        """
        canvas.scan_mark(event.x, event.y)

    def pan(event):
        """
        Handle panning the image.
        """
        canvas.scan_dragto(event.x, event.y, gain=1)

    # Create a new window to display the image
    window = Toplevel()
    window.title("Select Color from Image")

    # Create a frame for the canvas and scrollbars
    frame = tk.Frame(window)
    frame.pack(fill="both", expand=True)

    # Create a canvas to display the image
    canvas = Canvas(frame, bg="white")
    canvas.pack(side="left", fill="both", expand=True)

    # Add scrollbars
    h_scroll = tk.Scrollbar(frame, orient="horizontal", command=canvas.xview)
    h_scroll.pack(side="bottom", fill="x")
    v_scroll = tk.Scrollbar(frame, orient="vertical", command=canvas.yview)
    v_scroll.pack(side="right", fill="y")
    canvas.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)

    # Convert the image to a format suitable for Tkinter
    original_width, original_height = sample_image.size
    tk_image = ImageTk.PhotoImage(sample_image)
    scale = 1.0  # Initial zoom scale
    image_id = canvas.create_image(0, 0, anchor="nw", image=tk_image, tags="image")
    canvas.configure(scrollregion=canvas.bbox("all"))

    # Bind mouse events for color picking, zooming, and panning
    canvas.bind("<Button-1>", on_click)
    canvas.bind("<MouseWheel>", zoom)
    canvas.bind("<ButtonPress-2>", start_pan)  # Middle mouse button to start panning
    canvas.bind("<B2-Motion>", pan)  # Drag with middle mouse button to pan

    # Add a slider to adjust the tolerance value
    tolerance_slider = tk.Scale(window, from_=0, to=100, orient="horizontal", label="Tolerance", command=update_tolerance)
    tolerance_slider.set(tolerance_value)
    tolerance_slider.pack(fill="x", padx=10, pady=5)

    # Add a label to display the selected color
    color_label = Label(window, text="Selected Color: None", bg="white", font=("Arial", 12))
    color_label.pack(pady=10)

    # Keep a reference to the image to prevent garbage collection
    canvas.image = tk_image

def pick_color():
    """
    Open a color picker dialog to select the text color.
    """
    global selected_color
    color = askcolor(title="Select Text Color")[0]
    if color:
        selected_color = tuple(int(c) for c in color)
        print(f"Selected color: {selected_color}")

def create_export_to_excel_window():
    def toggle_filter():
        """
        Toggle the global apply_filter variable based on the checkbox state.
        """
        global apply_filter
        apply_filter = filter_checkbox_var.get()
        print(f"Apply filter: {apply_filter}")

    def browse_input_folder():
        """
        Open a folder dialog to select a custom input folder for images.
        """
        global cropped_images_folder
        folder_path = filedialog.askdirectory(title="Select Input Folder")
        if folder_path:
            cropped_images_folder = folder_path
            input_folder_label.config(text=cropped_images_folder)
            print(f"Selected input folder: {cropped_images_folder}")

    def browse_model_path():
        """
        Open a file dialog to select a model file.
        """
        global model_path
        file_path = filedialog.askopenfilename(filetypes=[("Model Files", "*.pth")], title="Select Model File")
        if file_path:
            model_path = file_path
            model_path_label.config(text=model_path)
            print(f"Selected model path: {model_path}")

    def update_margins():
        """
        Update the global margin values based on user input.
        """
        global margin_top, margin_bottom, margin_left, margin_right
        try:
            margin_top = int(margin_top_entry.get())
            margin_bottom = int(margin_bottom_entry.get())
            margin_left = int(margin_left_entry.get())
            margin_right = int(margin_right_entry.get())
            print(f"Updated margins - Top: {margin_top}, Bottom: {margin_bottom}, Left: {margin_left}, Right: {margin_right}")
        except ValueError:
            messagebox.showerror("Error", "Margins must be integers.")

    def on_export():
        allowed_chars = allowed_chars_entry.get()
        try:
            process_images(False, allowed_chars)
            messagebox.showinfo("Success", f"Excel file saved to: {output_excel_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    window = Toplevel()
    window.title("Export to Excel")

    # Model Path Selection
    Label(window, text="Model Path:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
    model_path_label = Label(window, text=model_path)
    model_path_label.grid(row=0, column=1, padx=10, pady=5, sticky="w")
    Button(window, text="Browse", command=browse_model_path).grid(row=0, column=2, padx=10, pady=5)

    Label(window, text="Cropped Images Folder:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
    input_folder_label = Label(window, text=cropped_images_folder)
    input_folder_label.grid(row=1, column=1, padx=10, pady=5, sticky="w")
    Button(window, text="Browse", command=browse_input_folder).grid(row=1, column=2, padx=10, pady=5)

    Label(window, text="Output Excel Path:").grid(row=2, column=0, padx=10, pady=5, sticky="w")
    Label(window, text=output_excel_path).grid(row=2, column=1, padx=10, pady=5, sticky="w")

    Label(window, text="Allowed Characters for OCR:").grid(row=3, column=0, padx=10, pady=10, sticky="w")
    allowed_chars_entry = Entry(window)
    allowed_chars_entry.insert(0, "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789.,!?;:'\"()[]{}<>@#$%^&*-_+=/\\|~` ")
    allowed_chars_entry.grid(row=3, column=1, padx=10, pady=10, sticky="w")

    filter_checkbox_var = tk.BooleanVar(value=apply_filter)
    filter_checkbox = tk.Checkbutton(window, text="Apply Filter", variable=filter_checkbox_var, command=toggle_filter)
    filter_checkbox.grid(row=4, column=0, columnspan=2, pady=5)

    Button(window, text="Browse Sample Image", command=open_sample_image).grid(row=5, column=0, columnspan=2, pady=5)
    Button(window, text="Pick Text Color", command=pick_color).grid(row=6, column=0, columnspan=2, pady=5)

    Label(window, text="Margins (pixels):").grid(row=7, column=0, padx=10, pady=5, sticky="w")
    Label(window, text="Top:").grid(row=7, column=1, padx=5, sticky="e")
    margin_top_entry = Entry(window, width=5)
    margin_top_entry.insert(0, str(margin_top))
    margin_top_entry.grid(row=7, column=2, padx=5, sticky="w")

    Label(window, text="Bottom:").grid(row=8, column=1, padx=5, sticky="e")
    margin_bottom_entry = Entry(window, width=5)
    margin_bottom_entry.insert(0, str(margin_bottom))
    margin_bottom_entry.grid(row=8, column=2, padx=5, sticky="w")

    Label(window, text="Left:").grid(row=9, column=1, padx=5, sticky="e")
    margin_left_entry = Entry(window, width=5)
    margin_left_entry.insert(0, str(margin_left))
    margin_left_entry.grid(row=9, column=2, padx=5, sticky="w")

    Label(window, text="Right:").grid(row=10, column=1, padx=5, sticky="e")
    margin_right_entry = Entry(window, width=5)
    margin_right_entry.insert(0, str(margin_right))
    margin_right_entry.grid(row=10, column=2, padx=5, sticky="w")

    Button(window, text="Update Margins", command=update_margins).grid(row=11, column=0, columnspan=3, pady=10)

    export_button = Button(window, text="Extract Text", command=on_export)
    export_button.grid(row=12, column=0, columnspan=2, pady=10)

def start_gui():
    root = tk.Tk()
    root.title("Main Application")
    Label(root, text="Main Application", font=("Arial", 16)).pack(pady=10)
    Button(root, text="Open Export to Excel", command=create_export_to_excel_window).pack(pady=5)

    root.mainloop()

if __name__ == "__main__":
    start_gui()
